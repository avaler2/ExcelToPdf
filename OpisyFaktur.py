import openpyxl
from datetime import datetime
import pandas as pd
import os
from pathlib import Path
import shutil

#Function to convert dates from Excel
def xldate_to_datetime(xldate):
	temp = datetime.datetime(1900, 1, 1)
	delta = datetime.timedelta(days=xldate)
	return temp+delta

#wypełnianie formatek w Excelu
def WypelnijFormatke(catalogue, lp, nrDokumentu, DataWystawienia, WydatkiOgolem, WydatkiKwalifikowalne, OpisUsługi, numerZadania, KategoriaKosztow, nrKsiegowy, DataZapłaty): 
    
    #otwórz formatkę opisu
    ExcelFormatka = openpyxl.load_workbook('FormatkaOpisu.xlsx')
    print(ExcelFormatka.sheetnames)
    sheetOpis = ExcelFormatka['Arkusz1']
    print('Nazwa arkusza: ', sheetOpis.title)
    
    sheetOpis['B5'] = nrDokumentu
    x = pd.to_datetime(DataWystawienia)
    sheetOpis['D5'] =  x
    sheetOpis['E9'] = WydatkiOgolem
    sheetOpis['E10'] = WydatkiKwalifikowalne
    sheetOpis['E13'] = OpisUsługi
    sheetOpis['E14'] = 'Zadanie ' + str(numerZadania)
    temp = str(KategoriaKosztow)
    ind = temp.find('-')
    kategKosztow = temp[0 : ind]
    sheetOpis['E15'] = kategKosztow
    nazwaKosztu = temp[(ind + 2): ]
    sheetOpis['E16'] = nazwaKosztu
    sheetOpis['E18'] = nrKsiegowy
    sheetOpis['C19'] = DataZapłaty

#create title and save
    titleTemp = 'opis_' + str(nrDokumentu)
    title = titleTemp.replace('/', '_')
    #path_parent = os.path.dirname(os.getcwd())
    #os.chdir(catalogue)
    ExcelFormatka.save('poz. ' + lp + '-' +title + '.xlsx')
    #os.chdir(path_parent)

def otworzExcela(fileName, catalogue):

    #otwórz WOP
    wop = openpyxl.load_workbook(fileName)
    print(wop.sheetnames)
    sheet = wop['Zestawienie Dokumentów']
    print('Nazwa arkusza: ', sheet.title)

    print('Analiza wierszy...')
    for row in range(4, sheet.max_row+1):
        lp = sheet['A' + str(row)].value
        numerZadania = sheet['B' + str(row)].value
        nrPartnera  = sheet['C' + str(row)].value
        nrDokumentu = sheet['D' + str(row)].value
        nrKsiegowy = sheet['E' + str(row)].value
        rodzajId = sheet['F' + str(row)].value
        NIPlubPesel = sheet['G' + str(row)].value
        DataWystawienia = sheet['H' + str(row)].value
        DataZapłaty = sheet['I' + str(row)].value
        DataDo = sheet['J' + str(row)].value
        DataOd = sheet['K' + str(row)].value
        OpisUsługi = sheet['L' + str(row)].value
        NrUmowy = sheet['M' + str(row)].value
        KwotaBrutto = sheet['N' + str(row)].value
        KwotaNetto = sheet['O' + str(row)].value
        FVKoryg = sheet['P' + str(row)].value
        Zalacznik = sheet['Q' + str(row)].value
        Uwagi = sheet['R' + str(row)].value
        KategoriaKosztow = sheet['S' + str(row)].value
        WydatkiOgolem = sheet['T' + str(row)].value
        WydatkiKwalifikowalne = sheet['U' + str(row)].value
        KwalifikowalnyVat = sheet['V' + str(row)].value
        Dofinansowanie = sheet['W' + str(row)].value
        Limity = sheet['X' + str(row)].value
        WydatkiLimit = sheet['Y' + str(row)].value

    # print(lp, KategoriaKosztow, KwotaBrutto, Dofinansowanie, OpisUsługi, nrDokumentu, DataWystawienia)

        WypelnijFormatke(catalogue, lp, nrDokumentu, DataWystawienia, WydatkiOgolem, WydatkiKwalifikowalne, OpisUsługi, numerZadania, KategoriaKosztow, nrKsiegowy, DataZapłaty)


#start
parent_dir = os.getcwd()
print(parent_dir)
#entries = Path('/Users/przemyslawpolanski/Dropbox (FREE)/!FREE/! SL2014 weryfikacja (KdP)/8. OpisyFV/')
    
entries = Path(parent_dir)
for entry in entries.iterdir():
    fileName = entry.name
    print(fileName)
    fileWithoutExtension = os.path.splitext(fileName)[0]

    print(fileName[0 : 3])
    if fileName[0 : 3] == 'WoP':

        #os.mkdir(fileWithoutExtension)
        #shutil.copy(fileName, fileWithoutExtension + '/' + fileName)
        os.chdir(fileName)
        otworzExcela(fileName, fileWithoutExtension)
        os.chdir(parent_dir)

